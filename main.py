import asyncio
import json
import os
import io
import csv
import time
import logging
from datetime import datetime, timezone, timedelta
from aiohttp import web, ClientSession, WSMsgType

import io as _io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
log = logging.getLogger(__name__)

trades = {}
funding = {}
positions = {}
market_map = {}
connected = False
last_update = 0
account_balance = {'value': None, 'field': ''}
initial_load_done = False
last_incremental = 0

TOKEN = os.environ.get('LIGHTER_TOKEN', '')
HL_WALLET = os.environ.get('HL_WALLET', '')  # Hyperliquid wallet address
HL_BASE = 'https://api.hyperliquid.xyz/info'

eur_cache = {}  # EUR/USD rate cache

# Aster state
aster_trades = {}
aster_funding = {}
aster_positions = {}
aster_account_value = None
aster_loading = False
aster_load_done = False
aster_last_update = 0

ASTER_BASE = 'https://fapi.asterdex.com'
ASTER_API_KEY = os.environ.get('ASTER_API_KEY', '')
ASTER_SECRET = os.environ.get('ASTER_SECRET', '')

# Hyperliquid state
hl_trades = {}
hl_funding = {}
hl_positions = {}
hl_account_value = None
hl_loading = False
hl_load_done = False
hl_last_update = 0
BASE = 'https://mainnet.zklighter.elliot.ai'
BASE_WS = 'wss://mainnet.zklighter.elliot.ai/stream'
GENESIS_MS = 1737072000000

def get_account():
    try: return TOKEN.split(':')[1]
    except: return None

def hdrs():
    return {'Authorization': TOKEN}

def to_ms(dt):
    return int(dt.timestamp() * 1000)

def from_ms(ms):
    return datetime.fromtimestamp(ms / 1000, tz=timezone.utc)

def today_start_ms():
    now = datetime.now(timezone.utc)
    return to_ms(now.replace(hour=0, minute=0, second=0, microsecond=0))

async def load_markets(session):
    global market_map
    try:
        async with session.get(BASE + '/api/v1/orderBookDetails') as r:
            if r.status == 200:
                for m in (await r.json()).get('order_book_details', []):
                    market_map[str(m['market_id'])] = m['symbol']
                log.info(f"Markets: {len(market_map)}")
    except Exception as e:
        log.error(f"Markets: {e}")

def sym(mid):
    return market_map.get(str(mid), f'market_{mid}')

def parse_trade_csv(text):
    result = {}
    try:
        for row in csv.DictReader(io.StringIO(text.strip())):
            market = row.get('Market', '?')
            side_raw = row.get('Side', '').lower()
            is_open = 'open' in side_raw
            is_long = 'long' in side_raw
            pnl_raw = row.get('Closed PnL', '-')
            pnl = None if pnl_raw in ('-', '', 'null') else float(pnl_raw)
            date_str = row.get('Date', '')
            try:
                ts = to_ms(datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S').replace(tzinfo=timezone.utc))
            except:
                ts = int(time.time() * 1000)
            price = float(row.get('Price', 0) or 0)
            size = float(row.get('Size', 0) or 0)
            fee = float(row.get('Fee', 0) or 0)
            tid = f"{date_str}_{market}_{side_raw}_{price}_{size}".replace(' ', '_')
            result[tid] = {
                'id': tid, 'symbol': market,
                'side': 'long' if is_long else 'short',
                'tradeType': 'open' if is_open else 'close',
                'price': price, 'size': size, 'pnl': pnl,
                'fee': fee, 'ts': ts, 'source': 'export'
            }
    except Exception as e:
        log.error(f"parse_trade_csv: {e}")
    return result

async def export_call(session, account, start_ms, end_ms, etype):
    url = f"{BASE}/api/v1/export?account_index={account}&type={etype}&start_timestamp={start_ms}&end_timestamp={end_ms}"
    try:
        async with session.get(url, headers=hdrs()) as r:
            if r.status != 200:
                return None
            data = await r.json()
            data_url = data.get('data_url') or data.get('url')
            if not data_url:
                return None
        async with session.get(data_url) as r:
            if r.status != 200:
                return None
            return await r.text()
    except Exception as e:
        log.debug(f"export_call {etype}: {e}")
        return None

async def load_all_funding(session, account, start_ts=None):
    now_ms = int(time.time() * 1000)
    start = start_ts or GENESIS_MS
    cursor = None
    total = 0
    while True:
        url = (f"{BASE}/api/v1/positionFunding"
               f"?account_index={account}&market_id=255&limit=100"
               f"&start_timestamp={start}&end_timestamp={now_ms}")
        if cursor:
            url += f"&cursor={cursor}"
        try:
            async with session.get(url, headers=hdrs()) as r:
                if r.status != 200:
                    break
                data = await r.json()
                items = data.get('position_fundings', [])
                if not items:
                    break
                for item in items:
                    fid = str(item.get('funding_id', ''))
                    mid = str(item.get('market_id', ''))
                    payment = float(item.get('change', 0))
                    ts_val = item.get('timestamp', now_ms)
                    if fid:
                        funding[fid] = {
                            'id': fid, 'symbol': sym(mid),
                            'side': item.get('position_side', 'long'),
                            'payment': payment, 'rate': item.get('rate', ''),
                            'ts': ts_val
                        }
                        total += 1
                cursor = data.get('next_cursor')
                if not cursor or len(items) < 100:
                    break
                await asyncio.sleep(0.2)
        except Exception as e:
            log.error(f"load_funding: {e}")
            break
    ft = round(sum(f['payment'] for f in funding.values()), 4)
    log.info(f"Funding: {total} payments, total={ft}")

async def historical_load(session, account):
    global initial_load_done
    log.info("=== HISTORICAL LOAD START ===")
    now = datetime.now(timezone.utc)
    genesis = from_ms(GENESIS_MS).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    cur = genesis
    chunks = []
    while cur < now:
        nxt = (cur.replace(day=28) + timedelta(days=4)).replace(day=1)
        chunks.append((to_ms(cur), to_ms(min(nxt, now))))
        cur = nxt
    log.info(f"Loading {len(chunks)} monthly chunks")
    for i, (s, e) in enumerate(chunks):
        label = from_ms(s).strftime('%Y-%m')
        text = await export_call(session, account, s, e, 'trade')
        if text:
            chunk = parse_trade_csv(text)
            trades.update(chunk)
            log.info(f"Chunk {i+1}/{len(chunks)} {label}: +{len(chunk)} trades (total {len(trades)})")
        await asyncio.sleep(0.3)
    await load_all_funding(session, account)
    wp = sum(1 for t in trades.values() if t.get('pnl') is not None)
    ft = round(sum(f['payment'] for f in funding.values()), 4)
    log.info(f"=== DONE: {len(trades)} trades ({wp} with PnL), {len(funding)} funding, total={ft} ===")
    initial_load_done = True

async def incremental_update(session, account):
    global last_incremental
    now_ms = int(time.time() * 1000)
    ts = today_start_ms()
    log.info("Incremental update...")
    text = await export_call(session, account, ts, now_ms, 'trade')
    if text:
        new = parse_trade_csv(text)
        before = len(trades)
        trades.update(new)
        log.info(f"Incremental: +{len(trades)-before} new trades")
    await asyncio.sleep(0.3)
    await load_all_funding(session, account, start_ts=ts)
    await load_positions(session, account)
    last_incremental = now_ms

async def load_positions(session, account):
    global positions, account_balance
    try:
        async with session.get(f"{BASE}/api/v1/account?by=index&value={account}", headers=hdrs()) as r:
            if r.status == 200:
                data = await r.json()
                # Extract account balance from accounts array
                accounts = data.get('accounts') or []
                if accounts:
                    acct = accounts[0]
                    # total_asset_value is the full account value including open positions
                    for key in ['total_asset_value','cross_asset_value','collateral','available_balance']:
                        val = acct.get(key)
                        if val is not None:
                            try:
                                account_balance['value'] = float(val)
                                account_balance['field'] = key
                                log.info(f"Account balance: {key}={val}")
                                break
                            except: pass
                    positions_data = acct.get('positions') or []
                else:
                    positions_data = (data.get('positions') or {})
                # handle both list and dict formats
                if isinstance(positions_data, list):
                    pos_iter = {str(p['market_id']): p for p in positions_data if float(p.get('position',0)) != 0}.items()
                else:
                    pos_iter = positions_data.items()
                for mid, pos in pos_iter:
                    positions[str(mid)] = {
                        'market_id': mid, 'symbol': sym(mid),
                        'side': 'long' if int(pos.get('sign', 1)) > 0 else 'short',
                        'size': float(pos.get('position', 0)),
                        'avg_entry': float(pos.get('avg_entry_price', 0)),
                        'unrealized_pnl': float(pos.get('unrealized_pnl', 0)),
                        'realized_pnl': float(pos.get('realized_pnl', 0)),
                        'liquidation_price': float(pos.get('liquidation_price', 0)),
                    }
    except Exception as e:
        log.error(f"positions: {e}")

def process_ws_trade(t, account):
    try:
        tid = str(t.get('trade_id') or t.get('id', ''))
        if not tid or tid in trades: return
        is_ask = str(t.get('ask_account_id', '')) == str(account)
        trades[tid] = {
            'id': tid, 'symbol': sym(str(t.get('market_id', ''))),
            'side': 'short' if is_ask else 'long',
            'tradeType': 'unknown',
            'price': float(t.get('price', 0)),
            'size': float(t.get('size', 0)),
            'pnl': None,
            'fee': float(t.get('taker_fee') or t.get('maker_fee') or 0),
            'ts': t.get('timestamp') or int(time.time() * 1000),
            'source': 'ws'
        }
        global last_update
        last_update = int(time.time() * 1000)
        log.info(f"WS trade: {trades[tid]['symbol']}")
    except Exception as e:
        log.debug(f"ws_trade: {e}")

def process_ws_position(p):
    try:
        mid = str(p.get('market_id', ''))
        if not mid: return
        positions[mid] = {
            'market_id': mid, 'symbol': sym(mid),
            'side': 'long' if int(p.get('sign', 1)) > 0 else 'short',
            'size': float(p.get('position', 0)),
            'avg_entry': float(p.get('avg_entry_price', 0)),
            'unrealized_pnl': float(p.get('unrealized_pnl', 0)),
            'realized_pnl': float(p.get('realized_pnl', 0)),
            'liquidation_price': float(p.get('liquidation_price', 0)),
        }
        global last_update
        last_update = int(time.time() * 1000)
    except Exception as e:
        log.debug(f"ws_pos: {e}")

async def hl_post(session, payload):
    try:
        async with session.post(HL_BASE,
                               json=payload,
                               headers={'Content-Type': 'application/json'}) as r:
            if r.status == 200:
                return await r.json()
    except Exception as e:
        log.error(f"HL API: {e}")
    return None

import hmac as _hmac
import hashlib as _hashlib

def aster_sign(params: dict) -> str:
    query = '&'.join(f"{k}={v}" for k, v in sorted(params.items()))
    return _hmac.new(ASTER_SECRET.encode(), query.encode(), _hashlib.sha256).hexdigest()

async def aster_get(session, path, params=None, signed=False):
    params = params or {}
    if signed:
        params['timestamp'] = int(time.time() * 1000)
        params['recvWindow'] = 10000
        params['signature'] = aster_sign(params)
    headers = {'X-MBX-APIKEY': ASTER_API_KEY}
    url = ASTER_BASE + path + ('?' + '&'.join(f"{k}={v}" for k, v in params.items()) if params else '')
    try:
        async with session.get(url, headers=headers) as r:
            if r.status == 200:
                return await r.json()
            elif r.status in (418, 429):
                body = await r.text()
                log.error(f"Aster {path} RATE LIMITED {r.status}: {body[:150]}")
                return None
            elif r.status == 401:
                body = await r.text()
                log.error(f"Aster {path} AUTH ERROR {r.status}: {body[:150]}")
                return None
            else:
                body = await r.text()
                log.error(f"Aster {path} HTTP {r.status}: {body[:150]}")
    except Exception as e:
        log.error(f"Aster {path}: {e}")
    return None

async def load_aster_income(session, income_type, start_ms):
    """Load income history with 7-day chunks (Aster limit)."""
    results = []
    now_ms = int(time.time() * 1000)
    chunk_ms = 7 * 24 * 60 * 60 * 1000  # 7 days max per request
    cursor = start_ms
    while cursor < now_ms:
        end = min(cursor + chunk_ms, now_ms)
        data = await aster_get(session, '/fapi/v1/income', {
            'incomeType': income_type,
            'startTime': cursor,
            'endTime': end,
            'limit': 1000
        }, signed=True)
        if data is None:
            # Error (401/418/429) — stop immediately
            log.warning(f"Aster {income_type}: stopping due to error")
            break
        if isinstance(data, list):
            results.extend(data)
            if len(data) > 0:
                log.info(f"Aster {income_type}: +{len(data)} (total {len(results)})")
        cursor = end + 1
        await asyncio.sleep(1.0)  # Be respectful with rate limits
    return results

async def prefetch_eur_rates():
    """Pre-fetch EUR rates for all dates we have trades on."""
    all_ts = (
        [int(t.get('ts',0)) for t in trades.values() if t.get('ts')] +
        [int(t.get('ts',0)) for t in hl_trades.values() if t.get('ts')] +
        [int(t.get('ts',0)) for t in aster_trades.values() if t.get('ts')]
    )
    dates = set(datetime.fromtimestamp(ts/1000, tz=timezone.utc).strftime('%Y-%m-%d') for ts in all_ts if ts)
    uncached = [d for d in dates if d not in eur_cache]
    if not uncached:
        return
    log.info(f"Fetching EUR rates for {len(uncached)} dates...")
    async with ClientSession() as session:
        for date_str in sorted(uncached):
            try:
                url = f"https://api.frankfurter.app/{date_str}?from=USD&to=EUR"
                async with session.get(url) as r:
                    if r.status == 200:
                        data = await r.json()
                        rate = data.get('rates', {}).get('EUR')
                        if rate:
                            eur_cache[date_str] = float(rate)
                await asyncio.sleep(0.1)
            except: pass
    log.info(f"EUR rates cached: {len(eur_cache)} dates")

async def load_aster_data():
    global aster_trades, aster_funding, aster_positions, aster_account_value
    global aster_loading, aster_load_done, aster_last_update
    if not ASTER_API_KEY or not ASTER_SECRET:
        log.info("No Aster credentials, skipping")
        return
    aster_loading = True
    log.info("Loading Aster data...")
    async with ClientSession() as session:
        # Account balance
        acct = await aster_get(session, '/fapi/v1/account', signed=True)
        if acct:
            aster_account_value = float(acct.get('totalWalletBalance') or acct.get('totalMarginBalance') or 0)
            log.info(f"Aster account value: {aster_account_value}")
            for pos in acct.get('positions', []):
                amt = float(pos.get('positionAmt', 0))
                if amt == 0:
                    continue
                symbol = pos.get('symbol', '?').replace('USDT', '').replace('USDC', '')
                aster_positions[symbol] = {
                    'symbol': symbol,
                    'side': 'long' if amt > 0 else 'short',
                    'size': abs(amt),
                    'avg_entry': float(pos.get('entryPrice', 0) or 0),
                    'unrealized_pnl': float(pos.get('unrealizedProfit', 0) or 0),
                    'realized_pnl': 0,
                    'liquidation_price': float(pos.get('liquidationPrice', 0) or 0),
                    'leverage': int(pos.get('leverage', 1) or 1),
                }

        # Trades with realized PnL — from income REALIZED_PNL
        # Aster genesis ~Jan 2024
        genesis_ms = 1704067200000
        pnl_rows = await load_aster_income(session, 'REALIZED_PNL', genesis_ms)
        for i, row in enumerate(pnl_rows):
            symbol = row.get('symbol', '?').replace('USDT', '').replace('USDC', '')
            pnl = float(row.get('income', 0))
            ts = int(row.get('time', 0))
            tid = f"at_{row.get('tranId', i)}_{ts}"
            aster_trades[tid] = {
                'id': tid, 'symbol': symbol,
                'side': 'long',  # not available in income endpoint
                'tradeType': 'close',
                'price': 0, 'size': 0,
                'pnl': pnl, 'fee': 0,
                'ts': ts, 'source': 'aster'
            }

        # Funding fees
        fund_rows = await load_aster_income(session, 'FUNDING_FEE', genesis_ms)
        for i, row in enumerate(fund_rows):
            symbol = row.get('symbol', '?').replace('USDT', '').replace('USDC', '')
            payment = float(row.get('income', 0))
            ts = int(row.get('time', 0))
            fid = f"af_{row.get('tranId', i)}_{ts}"
            aster_funding[fid] = {
                'id': fid, 'symbol': symbol,
                'payment': payment, 'ts': ts, 'source': 'aster'
            }

        aster_load_done = True
        aster_loading = False
        aster_last_update = int(time.time() * 1000)
        wp = len(aster_trades)
        ft = round(sum(f['payment'] for f in aster_funding.values()), 4)
        log.info(f"Aster DONE: {wp} trades with PnL, {len(aster_funding)} funding payments, funding={ft}")

async def aster_incremental(session=None):
    global aster_trades, aster_funding, aster_positions, aster_account_value, aster_last_update
    if not ASTER_API_KEY or not ASTER_SECRET or not aster_load_done:
        return
    log.info("Aster incremental update...")
    async with ClientSession() as s:
        ts = today_start_ms()
        now_ms = int(time.time() * 1000)
        # Today's PnL
        rows = await load_aster_income(s, 'REALIZED_PNL', ts)
        for i, row in enumerate(rows):
            symbol = row.get('symbol', '?').replace('USDT', '').replace('USDC', '')
            pnl = float(row.get('income', 0))
            ts2 = int(row.get('time', 0))
            tid = f"at_{row.get('tranId', i)}_{ts2}"
            aster_trades[tid] = {'id': tid, 'symbol': symbol, 'side': 'long',
                'tradeType': 'close', 'price': 0, 'size': 0,
                'pnl': pnl, 'fee': 0, 'ts': ts2, 'source': 'aster'}
        # Today's funding
        rows = await load_aster_income(s, 'FUNDING_FEE', ts)
        for i, row in enumerate(rows):
            symbol = row.get('symbol', '?').replace('USDT', '').replace('USDC', '')
            payment = float(row.get('income', 0))
            ts2 = int(row.get('time', 0))
            fid = f"af_{row.get('tranId', i)}_{ts2}"
            aster_funding[fid] = {'id': fid, 'symbol': symbol, 'payment': payment, 'ts': ts2, 'source': 'aster'}
        # Refresh positions and balance
        acct = await aster_get(s, '/fapi/v1/account', signed=True)
        if acct:
            aster_account_value = float(acct.get('totalWalletBalance') or acct.get('totalMarginBalance') or 0)
            aster_positions.clear()
            for pos in acct.get('positions', []):
                amt = float(pos.get('positionAmt', 0))
                if amt == 0: continue
                sym = pos.get('symbol', '?').replace('USDT', '').replace('USDC', '')
                aster_positions[sym] = {
                    'symbol': sym, 'side': 'long' if amt > 0 else 'short',
                    'size': abs(amt), 'avg_entry': float(pos.get('entryPrice', 0) or 0),
                    'unrealized_pnl': float(pos.get('unrealizedProfit', 0) or 0),
                    'realized_pnl': 0,
                    'liquidation_price': float(pos.get('liquidationPrice', 0) or 0),
                    'leverage': int(pos.get('leverage', 1) or 1),
                }
        aster_last_update = int(time.time() * 1000)
        log.info(f"Aster incremental done: {len(aster_trades)} trades, {len(aster_funding)} funding")

def build_aster_summary():
    closes = [t for t in aster_trades.values() if t.get('pnl') is not None]
    pnls = [t['pnl'] for t in closes]
    ft = round(sum(f['payment'] for f in aster_funding.values()), 4)
    tp = round(sum(pnls), 4)
    wins = sum(1 for p in pnls if p > 0)
    losses = sum(1 for p in pnls if p < 0)
    wr = round(wins / len(pnls) * 100, 1) if pnls else 0
    ts = today_start_ms()
    today_c = [t for t in closes if int(t.get('ts', 0) or 0) >= ts]
    today_pnl = round(sum(t['pnl'] for t in today_c), 4)
    today_f = round(sum(f['payment'] for f in aster_funding.values() if int(f.get('ts', 0) or 0) >= ts), 4)
    by_sym = {}
    for t in closes:
        s = t.get('symbol', '?')
        if s not in by_sym:
            by_sym[s] = {'symbol': s, 'trades': 0, 'pnl': 0.0, 'wins': 0, 'losses': 0, 'best': None, 'worst': None}
        m = by_sym[s]; m['trades'] += 1; m['pnl'] += t['pnl']
        if t['pnl'] > 0: m['wins'] += 1
        elif t['pnl'] < 0: m['losses'] += 1
        if m['best'] is None or t['pnl'] > m['best']: m['best'] = t['pnl']
        if m['worst'] is None or t['pnl'] < m['worst']: m['worst'] = t['pnl']
    for s in by_sym:
        by_sym[s]['pnl'] = round(by_sym[s]['pnl'], 4)
        if by_sym[s]['best']: by_sym[s]['best'] = round(by_sym[s]['best'], 4)
        if by_sym[s]['worst']: by_sym[s]['worst'] = round(by_sym[s]['worst'], 4)
    return {
        'total_pnl': round(tp + ft, 4), 'trade_pnl': tp, 'funding_total': ft,
        'today_pnl': round(today_pnl + today_f, 4), 'today_trades': len(today_c),
        'total_trades': len(aster_trades), 'closed_trades': len(closes),
        'wins': wins, 'losses': losses, 'win_rate': wr,
        'by_symbol': list(by_sym.values()),
        'positions': list(aster_positions.values()),
        'account_balance': aster_account_value,
        'loading': aster_loading, 'initial_load_done': aster_load_done,
        'last_update': aster_last_update
    }

async def load_hl_funding_all(session):
    """Load all funding using userNonFundingLedgerUpdates + userFundingHistory with pagination."""
    global hl_funding
    now_ms = int(time.time() * 1000)
    start_ms = 1700000000000
    total = 0

    # Method 1: userFundingHistory standard perps (paginated)
    cursor_ts = start_ms
    while True:
        data = await hl_post(session, {"type": "userFundingHistory", "user": HL_WALLET, "startTime": cursor_ts})
        if not data or not isinstance(data, list) or len(data) == 0:
            break
        for f in data:
            delta = f.get('delta', {})
            ts = int(f.get('time', 0))
            payment = float(delta.get('usdc', 0))
            coin = delta.get('coin', '?')
            fid = f"hl_std_{ts}_{coin}"
            hl_funding[fid] = {'id': fid, 'symbol': coin, 'payment': payment, 'ts': ts, 'source': 'hl_api'}
            total += 1
        if len(data) < 500:
            break
        cursor_ts = data[-1]['time'] + 1
        await asyncio.sleep(0.2)

    # Method 2: userNonFundingLedgerUpdates — contains funding-like entries for HIP-3
    cursor_ts = start_ms
    while True:
        data = await hl_post(session, {"type": "userNonFundingLedgerUpdates", "user": HL_WALLET, "startTime": cursor_ts})
        if not data or not isinstance(data, list) or len(data) == 0:
            break
        for f in data:
            delta = f.get('delta', {})
            dtype = delta.get('type', '')
            # Only include funding-type entries, not deposits/withdrawals/transfers
            if dtype in ('funding', 'fundingPayment', 'perpFunding', 'hip3Funding'):
                ts = int(f.get('time', 0))
                payment = float(delta.get('usdc', 0) or delta.get('amount', 0))
                coin = delta.get('coin', delta.get('asset', '?'))
                fid = f"hl_nfl_{ts}_{coin}_{dtype}"
                hl_funding[fid] = {'id': fid, 'symbol': coin, 'payment': payment, 'ts': ts, 'source': 'hl_ledger'}
                total += 1
        if len(data) < 500:
            break
        cursor_ts = data[-1]['time'] + 1
        await asyncio.sleep(0.2)

    ft = round(sum(f['payment'] for f in hl_funding.values()), 4)
    log.info(f"HL funding loaded: {len(hl_funding)} entries, total={ft} USDC")

async def load_hl_data():
    global hl_trades, hl_funding, hl_positions, hl_account_value, hl_loading, hl_load_done, hl_last_update
    if not HL_WALLET:
        log.info("No HL_WALLET set, skipping Hyperliquid")
        return
    hl_loading = True
    log.info(f"Loading Hyperliquid data for {HL_WALLET}...")
    async with ClientSession() as session:
        # 1. User fills (trades with closedPnl)
        data = await hl_post(session, {"type": "userFills", "user": HL_WALLET})
        if data and isinstance(data, list):
            for f in data:
                coin = f.get('coin', '?')
                dir_raw = (f.get('dir', '')).lower()
                is_open = 'open' in dir_raw
                is_long = 'long' in dir_raw or f.get('side','') == 'B'
                pnl_raw = f.get('closedPnl', '0')
                pnl = float(pnl_raw) if pnl_raw and pnl_raw != '0' and not is_open else None
                ts = int(f.get('time', 0))
                price = float(f.get('px', 0))
                size = float(f.get('sz', 0))
                fee = float(f.get('fee', 0))
                tid = str(f.get('tid', '')) or str(f.get('hash','')) or f"{ts}_{coin}_{price}"
                hl_trades[tid] = {
                    'id': tid,
                    'symbol': coin,
                    'side': 'long' if is_long else 'short',
                    'tradeType': 'open' if is_open else 'close',
                    'price': price,
                    'size': size,
                    'pnl': pnl if pnl != 0.0 else None,
                    'fee': fee,
                    'ts': ts,
                    'source': 'hl'
                }
            log.info(f"HL fills: {len(hl_trades)}")

        # 2. Funding via userNonFundingLedgerUpdates (paginated) + userFundingHistory standard
        await load_hl_funding_all(session)

        # 3. Current positions + account value from portfolio + spot balance
        data = await hl_post(session, {"type": "clearinghouseState", "user": HL_WALLET})
        if data:
            for pos in (data.get('assetPositions', [])):
                p = pos.get('position', {})
                coin = p.get('coin', '?')
                size = float(p.get('szi', 0))
                if size == 0:
                    continue
                hl_positions[coin] = {
                    'symbol': coin,
                    'side': 'long' if size > 0 else 'short',
                    'size': abs(size),
                    'avg_entry': float(p.get('entryPx', 0) or 0),
                    'unrealized_pnl': float(p.get('unrealizedPnl', 0) or 0),
                    'realized_pnl': 0,
                    'liquidation_price': float(p.get('liquidationPx', 0) or 0),
                    'leverage': p.get('leverage', {}).get('value', 1),
                }

        # Also load HIP-3 positions (dex=xyz)
        xyz_data = await hl_post(session, {"type": "clearinghouseState", "user": HL_WALLET, "dex": "xyz"})
        if xyz_data:
            xyz_margin = float(xyz_data.get('marginSummary', {}).get('accountValue', 0))
            for pos in (xyz_data.get('assetPositions', [])):
                p = pos.get('position', {})
                coin = 'xyz:' + p.get('coin', '?')
                size = float(p.get('szi', 0))
                if size == 0: continue
                hl_positions[coin] = {
                    'symbol': coin,
                    'side': 'long' if size > 0 else 'short',
                    'size': abs(size),
                    'avg_entry': float(p.get('entryPx', 0) or 0),
                    'unrealized_pnl': float(p.get('unrealizedPnl', 0) or 0),
                    'realized_pnl': 0,
                    'liquidation_price': float(p.get('liquidationPx', 0) or 0),
                    'leverage': p.get('leverage', {}).get('value', 1),
                }

        # Account value = spot USDC balance + HIP-3 perp account value
        spot = await hl_post(session, {"type": "spotClearinghouseState", "user": HL_WALLET})
        spot_usdc = 0
        if spot:
            for bal in spot.get('balances', []):
                if bal.get('coin') == 'USDC':
                    spot_usdc = float(bal.get('total', 0))
                    break
        xyz_val = float(xyz_data.get('marginSummary', {}).get('accountValue', 0)) if xyz_data else 0
        hl_account_value = spot_usdc + xyz_val

        log.info(f"HL positions: {len(hl_positions)}, account value: {hl_account_value}")

        hl_load_done = True
        hl_loading = False
        hl_last_update = int(time.time() * 1000)
        wp = sum(1 for t in hl_trades.values() if t.get('pnl') is not None)
        ft = round(sum(f['payment'] for f in hl_funding.values()), 4)
        log.info(f"HL DONE: {len(hl_trades)} fills ({wp} with PnL), funding={ft}")

async def hl_incremental():
    global hl_trades, hl_positions, hl_account_value, hl_funding, hl_last_update
    if not HL_WALLET or not hl_load_done:
        return
    log.info("HL incremental update...")
    async with ClientSession() as session:
        await load_hl_funding_all(session)
        # Get latest fills
        data = await hl_post(session, {"type": "userFills", "user": HL_WALLET})
        if data and isinstance(data, list):
            before = len(hl_trades)
            for f in data:
                coin = f.get('coin', '?')
                dir_raw = (f.get('dir', '')).lower()
                is_open = 'open' in dir_raw
                is_long = 'long' in dir_raw or f.get('side','') == 'B'
                pnl_raw = f.get('closedPnl', '0')
                pnl = float(pnl_raw) if pnl_raw and pnl_raw != '0' and not is_open else None
                ts = int(f.get('time', 0))
                price = float(f.get('px', 0))
                size = float(f.get('sz', 0))
                fee = float(f.get('fee', 0))
                tid = str(f.get('tid', '')) or f"{ts}_{coin}_{price}"
                hl_trades[tid] = {
                    'id': tid, 'symbol': coin,
                    'side': 'long' if is_long else 'short',
                    'tradeType': 'open' if is_open else 'close',
                    'price': price, 'size': size,
                    'pnl': pnl if pnl != 0.0 else None,
                    'fee': fee, 'ts': ts, 'source': 'hl'
                }
            log.info(f"HL incremental: +{len(hl_trades)-before} new fills")
        # Refresh positions and account value
        data = await hl_post(session, {"type": "clearinghouseState", "user": HL_WALLET})
        xyz_data2 = await hl_post(session, {"type": "clearinghouseState", "user": HL_WALLET, "dex": "xyz"})
        if data:
            hl_positions.clear()
            for pos in (data.get('assetPositions', [])):
                p = pos.get('position', {})
                coin = p.get('coin', '?')
                size = float(p.get('szi', 0))
                if size == 0: continue
                hl_positions[coin] = {
                    'symbol': coin, 'side': 'long' if size > 0 else 'short',
                    'size': abs(size), 'avg_entry': float(p.get('entryPx', 0) or 0),
                    'unrealized_pnl': float(p.get('unrealizedPnl', 0) or 0),
                    'realized_pnl': 0,
                    'liquidation_price': float(p.get('liquidationPx', 0) or 0),
                    'leverage': p.get('leverage', {}).get('value', 1),
                }
        if xyz_data2:
            for pos in (xyz_data2.get('assetPositions', [])):
                p = pos.get('position', {})
                coin = 'xyz:' + p.get('coin', '?')
                size = float(p.get('szi', 0))
                if size == 0: continue
                hl_positions[coin] = {
                    'symbol': coin, 'side': 'long' if size > 0 else 'short',
                    'size': abs(size), 'avg_entry': float(p.get('entryPx', 0) or 0),
                    'unrealized_pnl': float(p.get('unrealizedPnl', 0) or 0),
                    'realized_pnl': 0,
                    'liquidation_price': float(p.get('liquidationPx', 0) or 0),
                    'leverage': p.get('leverage', {}).get('value', 1),
                }
        # Recalculate account value
        spot2 = await hl_post(session, {"type": "spotClearinghouseState", "user": HL_WALLET})
        spot_usdc2 = 0
        if spot2:
            for bal in spot2.get('balances', []):
                if bal.get('coin') == 'USDC':
                    spot_usdc2 = float(bal.get('total', 0))
                    break
        xyz_val2 = float(xyz_data2.get('marginSummary', {}).get('accountValue', 0)) if xyz_data2 else 0
        hl_account_value = spot_usdc2 + xyz_val2
        hl_last_update = int(time.time() * 1000)

def build_hl_summary():
    closes = [t for t in hl_trades.values() if t.get('tradeType') == 'close' and t.get('pnl') is not None]
    pnls = [t['pnl'] for t in closes]
    ft = round(sum(f['payment'] for f in hl_funding.values()), 4)
    tp = round(sum(pnls), 4)
    wins = sum(1 for p in pnls if p > 0)
    losses = sum(1 for p in pnls if p < 0)
    wr = round(wins / len(pnls) * 100, 1) if pnls else 0
    ts = today_start_ms()
    today_c = [t for t in closes if int(t.get('ts', 0) or 0) >= ts]
    today_pnl = round(sum(t['pnl'] for t in today_c), 4)
    today_f = round(sum(f['payment'] for f in hl_funding.values() if int(f.get('ts',0) or 0) >= ts), 4)
    by_sym = {}
    for t in closes:
        s = t.get('symbol', '?')
        if s not in by_sym:
            by_sym[s] = {'symbol': s, 'trades': 0, 'pnl': 0.0, 'wins': 0, 'losses': 0, 'best': None, 'worst': None}
        m = by_sym[s]; m['trades'] += 1; m['pnl'] += t['pnl']
        if t['pnl'] > 0: m['wins'] += 1
        elif t['pnl'] < 0: m['losses'] += 1
        if m['best'] is None or t['pnl'] > m['best']: m['best'] = t['pnl']
        if m['worst'] is None or t['pnl'] < m['worst']: m['worst'] = t['pnl']
    for s in by_sym:
        by_sym[s]['pnl'] = round(by_sym[s]['pnl'], 4)
        if by_sym[s]['best']: by_sym[s]['best'] = round(by_sym[s]['best'], 4)
        if by_sym[s]['worst']: by_sym[s]['worst'] = round(by_sym[s]['worst'], 4)
    return {
        'total_pnl': round(tp + ft, 4),
        'trade_pnl': tp, 'funding_total': ft,
        'today_pnl': round(today_pnl + today_f, 4),
        'today_trades': len(today_c),
        'total_trades': len(hl_trades), 'closed_trades': len(closes),
        'wins': wins, 'losses': losses, 'win_rate': wr,
        'by_symbol': list(by_sym.values()),
        'positions': list(hl_positions.values()),
        'account_balance': hl_account_value,
        'loading': hl_loading,
        'initial_load_done': hl_load_done,
        'last_update': hl_last_update
    }

async def ws_listener():
    global connected, last_update
    account = get_account()
    if not TOKEN or not account:
        log.error("No LIGHTER_TOKEN")
        return
    async with ClientSession() as session:
        await load_markets(session)
        await load_positions(session, account)
        await historical_load(session, account)
        async def scheduler():
            while True:
                await asyncio.sleep(900)
                await incremental_update(session, account)
        asyncio.ensure_future(scheduler())
        while True:
            try:
                async with session.ws_connect(BASE_WS, heartbeat=60) as ws:
                    connected = True
                    await ws.send_json({"type": "subscribe", "channel": f"account_all_trades/{account}", "auth": TOKEN})
                    await ws.send_json({"type": "subscribe", "channel": f"account_all_positions/{account}", "auth": TOKEN})
                    log.info(f"WS connected account {account}")
                    async for msg in ws:
                        if msg.type == WSMsgType.TEXT:
                            try:
                                d = json.loads(msg.data)
                                mt = d.get('type', '')
                                if 'trade' in mt.lower():
                                    td = d.get('trade') or d.get('trades') or d.get('data')
                                    for t in ([td] if isinstance(td, dict) else (td or [])):
                                        process_ws_trade(t, account)
                                elif 'position' in mt.lower():
                                    pd = d.get('position') or d.get('positions') or d.get('data')
                                    for p in ([pd] if isinstance(pd, dict) else (pd or [])):
                                        process_ws_position(p)
                                last_update = int(time.time() * 1000)
                            except: pass
                        elif msg.type in (WSMsgType.CLOSED, WSMsgType.ERROR):
                            break
            except Exception as e:
                log.error(f"WS: {e}")
            connected = False
            await asyncio.sleep(5)

def cors(r):
    r.headers['Access-Control-Allow-Origin'] = '*'
    return r

async def h_root(req):
    return cors(web.json_response({'ok': True, 'loading': not initial_load_done}))

async def h_status(req):
    tp = round(sum(t['pnl'] for t in trades.values() if t.get('pnl') is not None), 4)
    ft = round(sum(f['payment'] for f in funding.values()), 4)
    return cors(web.json_response({
        'ok': True, 'connected': connected,
        'account': get_account(),
        'initial_load_done': initial_load_done,
        'trades': len(trades), 'funding': len(funding),
        'positions': len(positions),
        'trade_pnl': tp, 'funding_total': ft,
        'total_pnl': round(tp + ft, 4),
        'last_incremental': last_incremental,
        'last_update': last_update,
        'ts': int(time.time() * 1000)
    }))

async def h_trades(req):
    limit = int(req.rel_url.query.get('limit', 20000))
    sym_f = req.rel_url.query.get('symbol', '').lower()
    all_t = sorted(trades.values(), key=lambda t: int(t.get('ts', 0) or 0), reverse=True)
    if sym_f:
        all_t = [t for t in all_t if sym_f in (t.get('symbol') or '').lower()]
    return cors(web.json_response({
        'trades': all_t[:limit],
        'total': len(all_t),
        'loading': not initial_load_done
    }))

async def h_funding(req):
    all_f = sorted(funding.values(), key=lambda f: int(f.get('ts', 0) or 0), reverse=True)
    return cors(web.json_response({
        'funding': all_f,
        'total': round(sum(f['payment'] for f in all_f), 4),
        'count': len(all_f)
    }))

async def h_positions(req):
    return cors(web.json_response({'positions': list(positions.values())}))

async def h_eur_rates(req):
    """EUR rates endpoint - pending implementation."""
    return cors(web.json_response({'current_rate': None, 'cached_dates': 0, 'note': 'EUR conversion coming soon'}))

async def h_summary(req):
    ts = today_start_ms()
    closes = [t for t in trades.values() if t.get('tradeType') == 'close' and t.get('pnl') is not None]
    pnls = [t['pnl'] for t in closes]
    tp = round(sum(pnls), 4)
    ft = round(sum(f['payment'] for f in funding.values()), 4)
    wins = sum(1 for p in pnls if p > 0)
    losses = sum(1 for p in pnls if p < 0)
    wr = round(wins / len(pnls) * 100, 1) if pnls else 0
    today_c = [t for t in closes if int(t.get('ts', 0) or 0) >= ts]
    today_pnl = round(sum(t['pnl'] for t in today_c), 4)
    today_f = round(sum(f['payment'] for f in funding.values() if int(f.get('ts', 0) or 0) >= ts), 4)
    by_sym = {}
    for t in closes:
        s = t.get('symbol', '?')
        if s not in by_sym:
            by_sym[s] = {'symbol': s, 'trades': 0, 'pnl': 0.0, 'wins': 0, 'losses': 0, 'best': None, 'worst': None}
        m = by_sym[s]
        m['trades'] += 1; m['pnl'] += t['pnl']
        if t['pnl'] > 0: m['wins'] += 1
        elif t['pnl'] < 0: m['losses'] += 1
        if m['best'] is None or t['pnl'] > m['best']: m['best'] = t['pnl']
        if m['worst'] is None or t['pnl'] < m['worst']: m['worst'] = t['pnl']
    for s in by_sym:
        by_sym[s]['pnl'] = round(by_sym[s]['pnl'], 4)
        if by_sym[s]['best']: by_sym[s]['best'] = round(by_sym[s]['best'], 4)
        if by_sym[s]['worst']: by_sym[s]['worst'] = round(by_sym[s]['worst'], 4)
    return cors(web.json_response({
        'total_pnl': round(tp + ft, 4),
        'trade_pnl': tp, 'funding_total': ft,
        'today_pnl': round(today_pnl + today_f, 4),
        'today_trade_pnl': today_pnl, 'today_funding': today_f,
        'total_trades': len(trades), 'closed_trades': len(closes),
        'today_trades': len(today_c), 'wins': wins, 'losses': losses, 'win_rate': wr,
        'by_symbol': list(by_sym.values()),
        'positions': list(positions.values()),
        'connected': connected,
        'initial_load_done': initial_load_done,
        'account_balance': account_balance.get('value'),
        'last_update': last_update
    }))

async def h_hl_upload_funding(req):
    """Accept CSV funding data uploaded from frontend."""
    global hl_funding
    try:
        data = await req.json()
        rows = data.get('rows', [])
        added = 0
        for i, row in enumerate(rows):
            # Expected fields: time/hora, coin/moneda, amount/pago, rate/tasa, side, size
            ts_raw = row.get('time') or row.get('hora') or row.get('Time') or row.get('Hora') or ''
            coin = row.get('coin') or row.get('moneda') or row.get('Coin') or row.get('Moneda') or '?'
            payment_raw = row.get('amount') or row.get('pago') or row.get('Amount') or row.get('Pago') or row.get('payment') or '0'
            # Parse timestamp
            try:
                if isinstance(ts_raw, (int, float)):
                    ts = int(ts_raw)
                else:
                    from datetime import datetime, timezone
                    ts_str = str(ts_raw).replace('/', '-')
                    dt = datetime.fromisoformat(ts_str.replace(' ', 'T'))
                    if dt.tzinfo is None:
                        dt = dt.replace(tzinfo=timezone.utc)
                    ts = int(dt.timestamp() * 1000)
            except:
                ts = int(time.time() * 1000)
            payment = float(str(payment_raw).replace(',', '.').replace(' USDC', '') or 0)
            fid = f"hl_csv_{ts}_{coin}_{i}"
            hl_funding[fid] = {
                'id': fid,
                'symbol': coin,
                'payment': payment,
                'ts': ts,
                'source': 'hl_csv'
            }
            added += 1
        total = round(sum(f['payment'] for f in hl_funding.values()), 4)
        log.info(f"HL funding CSV upload: +{added} rows, total={total}")
        return cors(web.json_response({'ok': True, 'added': added, 'total_funding': total, 'total_rows': len(hl_funding)}))
    except Exception as e:
        return cors(web.json_response({'error': str(e)}, status=400))

async def h_hl_clear_funding(req):
    """Clear all CSV-uploaded funding data."""
    global hl_funding
    hl_funding = {k: v for k, v in hl_funding.items() if v.get('source') != 'hl_csv'}
    return cors(web.json_response({'ok': True, 'remaining': len(hl_funding)}))

async def h_hl_ledger_inspect(req):
    wallet = HL_WALLET
    async with ClientSession() as session:
        data = await hl_post(session, {"type": "userNonFundingLedgerUpdates", "user": wallet, "startTime": 1700000000000})
        # Show all unique delta types and all entries
        types = {}
        for f in (data or []):
            dtype = f.get('delta', {}).get('type', 'unknown')
            if dtype not in types:
                types[dtype] = f
        return cors(web.json_response({
            'total': len(data) if data else 0,
            'unique_types': list(types.keys()),
            'examples': types
        }))

async def h_hl_funding_test(req):
    """Test all possible funding endpoints with pagination."""
    wallet = HL_WALLET
    results = {}
    async with ClientSession() as session:
        # Test every dex name variant
        for dex in ['', 'xyz', 'XYZ', 'hip3', 'HIP3']:
            payload = {"type": "userFundingHistory", "user": wallet, "startTime": 1700000000000}
            if dex:
                payload["dex"] = dex
            data = await hl_post(session, payload)
            key = f"dex_{dex or 'none'}"
            results[key] = {
                'count': len(data) if data and isinstance(data, list) else 0,
                'first': data[0] if data and isinstance(data, list) and len(data) > 0 else None,
                'raw_type': type(data).__name__,
                'raw_preview': str(data)[:200] if data else None
            }

        # Also try userNonFundingLedgerUpdates
        for type_name in ['userNonFundingLedgerUpdates', 'userRateLimit']:
            try:
                data = await hl_post(session, {"type": type_name, "user": wallet, "startTime": 1700000000000})
                results[type_name] = {
                    'count': len(data) if isinstance(data, list) else 'not_list',
                    'first': data[0] if isinstance(data, list) and data else None
                }
            except Exception as e:
                results[type_name] = str(e)

        # Try clearinghouseState with dex=xyz
        try:
            data = await hl_post(session, {"type": "clearinghouseState", "user": wallet, "dex": "xyz"})
            results['clearinghouseState_xyz'] = {
                'marginSummary': data.get('marginSummary') if data else None,
                'positions_count': len(data.get('assetPositions', [])) if data else 0
            }
        except Exception as e:
            results['clearinghouseState_xyz'] = str(e)

    return cors(web.json_response(results))

async def h_hl_debug(req):
    wallet = HL_WALLET
    async with ClientSession() as session:
        results = {}
        # Try multiple endpoints to find balance
        for etype in [
            {"type": "clearinghouseState", "user": wallet},
            {"type": "spotClearinghouseState", "user": wallet},
            {"type": "portfolio", "user": wallet},
            {"type": "userState", "user": wallet},
        ]:
            try:
                data = await hl_post(session, etype)
                results[etype['type']] = data
            except Exception as e:
                results[etype['type']] = str(e)

        # Try all funding endpoints
        for ftype in [
            {"type": "userFundingHistory", "user": wallet, "startTime": 1700000000000},
            {"type": "userFundingHistory", "user": wallet, "startTime": 1748000000000},
            {"type": "fundingHistory", "coin": "HYPE", "startTime": 1700000000000},
        ]:
            try:
                fd = await hl_post(session, ftype)
                results[str(ftype)] = {'count': len(fd) if fd else 0, 'first': fd[0] if fd else None}
            except Exception as e:
                results[str(ftype)] = str(e)

        # Try portfolio pnlHistory allTime as funding proxy
        port = await hl_post(session, {"type": "portfolio", "user": wallet})
        if port:
            for p in port:
                if p[0] == 'allTime':
                    results['portfolio_allTime_pnl_last'] = p[1].get('pnlHistory', [])[-3:]
                    results['portfolio_allTime_acv_last'] = p[1].get('accountValueHistory', [])[-1]

        return cors(web.json_response(results))

async def h_aster_upload_csv(req):
    global aster_trades, aster_funding
    try:
        data = await req.json()
        rows = data.get('rows', [])
        trades_added = 0
        funding_added = 0
        for i, row in enumerate(rows):
            time_str = row.get('Time', '') or row.get('time', '')
            rtype = row.get('Type', '') or row.get('type', '')
            amount_raw = row.get('Amount', '') or row.get('amount', '0')
            symbol = row.get('Symbol', '') or row.get('symbol', '')
            # Parse amount - remove currency suffix
            amount_clean = amount_raw.split(' ')[0] if amount_raw else '0'
            try:
                amount = float(amount_clean)
            except:
                amount = 0.0
            # Parse timestamp
            try:
                from datetime import datetime, timezone
                dt = datetime.strptime(time_str, '%Y-%m-%d %H:%M:%S').replace(tzinfo=timezone.utc)
                ts = int(dt.timestamp() * 1000)
            except:
                ts = int(time.time() * 1000)
            # Clean symbol
            sym = symbol.replace('USDT','').replace('USDC','').strip()
            rtype_lower = rtype.lower()
            if 'pnl' in rtype_lower or 'realizado' in rtype_lower or 'realized' in rtype_lower:
                tid = f"at_csv_{ts}_{sym}_{i}"
                aster_trades[tid] = {
                    'id': tid, 'symbol': sym or '?',
                    'side': 'long', 'tradeType': 'close',
                    'price': 0, 'size': 0,
                    'pnl': amount, 'fee': 0,
                    'ts': ts, 'source': 'aster_csv'
                }
                trades_added += 1
            elif 'financiamiento' in rtype_lower or 'funding' in rtype_lower:
                fid = f"af_csv_{ts}_{sym}_{i}"
                aster_funding[fid] = {
                    'id': fid, 'symbol': sym or '?',
                    'payment': amount, 'ts': ts, 'source': 'aster_csv'
                }
                funding_added += 1
        trade_pnl = round(sum(t['pnl'] for t in aster_trades.values() if t.get('pnl')), 4)
        fund_total = round(sum(f['payment'] for f in aster_funding.values()), 4)
        aster_load_done = True
        log.info(f"Aster CSV: +{trades_added} trades, +{funding_added} funding. PnL={trade_pnl}, Funding={fund_total}")
        return cors(web.json_response({
            'ok': True, 'trades_added': trades_added,
            'funding_added': funding_added,
            'trade_pnl': trade_pnl, 'funding_total': fund_total
        }))
    except Exception as e:
        return cors(web.json_response({'error': str(e)}, status=400))

async def h_aster_clear_csv(req):
    global aster_trades, aster_funding
    aster_trades = {k:v for k,v in aster_trades.items() if v.get('source') != 'aster_csv'}
    aster_funding = {k:v for k,v in aster_funding.items() if v.get('source') != 'aster_csv'}
    return cors(web.json_response({'ok': True}))

async def h_aster_summary(req):
    return cors(web.json_response(build_aster_summary()))

async def h_aster_trades(req):
    limit = int(req.rel_url.query.get('limit', 20000))
    all_t = sorted(aster_trades.values(), key=lambda t: int(t.get('ts', 0) or 0), reverse=True)
    return cors(web.json_response({'trades': all_t[:limit], 'total': len(all_t)}))

async def h_aster_positions(req):
    return cors(web.json_response({'positions': list(aster_positions.values())}))

async def h_hl_summary(req):
    return cors(web.json_response(build_hl_summary()))

async def h_hl_trades(req):
    limit = int(req.rel_url.query.get('limit', 20000))
    all_t = sorted(hl_trades.values(), key=lambda t: int(t.get('ts',0) or 0), reverse=True)
    return cors(web.json_response({'trades': all_t[:limit], 'total': len(all_t)}))

async def h_hl_positions(req):
    return cors(web.json_response({'positions': list(hl_positions.values())}))

async def h_account_debug(req):
    account = get_account()
    try:
        async with ClientSession() as s:
            async with s.get(f"{BASE}/api/v1/account?by=index&value={account}", headers=hdrs()) as r:
                data = await r.json()
                return cors(web.json_response({'raw': data}))
    except Exception as e:
        return cors(web.json_response({'error': str(e)}))

async def h_export_excel(req):
    """Generate Excel with all trading data across all exchanges."""
    try:
        import io as _io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        wb.remove(wb.active)

        # Styles
        hdr_font = Font(bold=True, color='FFFFFF', size=11)
        hdr_fill_lt = PatternFill('solid', fgColor='1a1a18')
        hdr_fill_hl = PatternFill('solid', fgColor='1D9E75')
        hdr_fill_at = PatternFill('solid', fgColor='A0522D')
        hdr_fill_gl = PatternFill('solid', fgColor='2C3E50')
        hdr_fill_sub = PatternFill('solid', fgColor='4A4A48')
        pos_fill = PatternFill('solid', fgColor='E8F5E9')
        neg_fill = PatternFill('solid', fgColor='FFEBEE')
        pos_font = Font(color='1B5E20', bold=True)
        neg_font = Font(color='B71C1C', bold=True)
        thin = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        center = Alignment(horizontal='center', vertical='center')
        left = Alignment(horizontal='left', vertical='center')

        def style_header_row(ws, row_num, headers, fill):
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row_num, col, h)
                cell.font = hdr_font; cell.fill = fill
                cell.alignment = center; cell.border = thin

        def set_col_widths(ws, widths):
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

        def style_pnl_cell(cell, val):
            if val is None or val == '': return
            try:
                v = float(str(val).replace('%','').replace('+',''))
                if v > 0: cell.font = pos_font; cell.fill = pos_fill
                elif v < 0: cell.font = neg_font; cell.fill = neg_fill
            except: pass

        def fmt_ts(ts):
            if not ts: return ''
            try: return datetime.fromtimestamp(int(ts)/1000, tz=timezone.utc).strftime('%Y-%m-%d %H:%M')
            except: return str(ts)

        def fmt_date(ts):
            if not ts: return ''
            try: return datetime.fromtimestamp(int(ts)/1000, tz=timezone.utc).strftime('%Y-%m-%d')
            except: return str(ts)

        def write_summary_block(ws, title, rows, fill, start_row=1):
            # Title
            tc = ws.cell(start_row, 1, title)
            tc.font = Font(bold=True, color='FFFFFF', size=13)
            tc.fill = fill; tc.alignment = left
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=3)
            r = start_row + 1
            for key, val in rows:
                ws.cell(r, 1, key).font = Font(bold=True)
                vc = ws.cell(r, 2, val)
                style_pnl_cell(vc, val)
                r += 1
            return r + 1

        def build_daily_map(trade_list):
            dm = {}
            for t in trade_list:
                if t.get('pnl') is None: continue
                k = fmt_date(t.get('ts',''))
                if k: dm[k] = dm.get(k, 0) + t['pnl']
            return dm

        # ── Prepare data ──
        lt_closes = [t for t in trades.values() if t.get('tradeType')=='close' and t.get('pnl') is not None]
        lt_pnl = sum(t['pnl'] for t in lt_closes)
        lt_fund = sum(f['payment'] for f in funding.values())
        lt_wins = sum(1 for t in lt_closes if t['pnl']>0)
        lt_wr = round(lt_wins/len(lt_closes)*100,1) if lt_closes else 0
        lt_bal = account_balance.get('value') or 0

        hl_closes = [t for t in hl_trades.values() if t.get('tradeType')=='close' and t.get('pnl') is not None]
        hl_pnl = sum(t['pnl'] for t in hl_closes)
        hl_fund = sum(f['payment'] for f in hl_funding.values())
        hl_wins = sum(1 for t in hl_closes if t['pnl']>0)
        hl_wr = round(hl_wins/len(hl_closes)*100,1) if hl_closes else 0
        hl_bal = hl_account_value or 0

        at_closes = [t for t in aster_trades.values() if t.get('pnl') is not None]
        at_pnl = sum(t['pnl'] for t in at_closes)
        at_fund = sum(f['payment'] for f in aster_funding.values())
        at_wins = sum(1 for t in at_closes if t['pnl']>0)
        at_wr = round(at_wins/len(at_closes)*100,1) if at_closes else 0
        at_bal = aster_account_value or 0

        total_pnl = lt_pnl + hl_pnl + at_pnl
        total_fund = lt_fund + hl_fund + at_fund
        total_bal = lt_bal + hl_bal + at_bal
        all_closes_n = len(lt_closes) + len(hl_closes) + len(at_closes)
        all_wins_n = lt_wins + hl_wins + at_wins
        global_wr = round(all_wins_n/all_closes_n*100,1) if all_closes_n else 0
        pct_rent = round((total_pnl+total_fund)/total_bal*100,2) if total_bal else 0

        # ══════════════════════════════════════════
        # HOJA GLOBAL (primera)
        # ══════════════════════════════════════════
        ws_gl = wb.create_sheet('Global', 0)

        # Title block
        ws_gl.row_dimensions[1].height = 30
        t1 = ws_gl.cell(1, 1, 'REGISTRO DE OPERACIONES DEX')
        t1.font = Font(bold=True, size=16, color='FFFFFF')
        t1.fill = hdr_fill_gl; t1.alignment = center
        ws_gl.merge_cells('A1:E1')
        ws_gl.cell(2, 1, f"Generado: {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M UTC')}").font = Font(italic=True, color='888888')
        ws_gl.merge_cells('A2:E2')

        # Resumen por exchange
        ws_gl.cell(4, 1, 'RESUMEN POR EXCHANGE').font = Font(bold=True, size=13, color='FFFFFF')
        ws_gl.cell(4, 1).fill = hdr_fill_gl
        ws_gl.merge_cells('A4:E4')
        style_header_row(ws_gl, 5, ['', 'Lighter', 'Hyperliquid', 'Aster', 'TOTAL'], hdr_fill_sub)
        summary_data = [
            ('P&L Trades (USDC)', round(lt_pnl,4), round(hl_pnl,4), round(at_pnl,4), round(total_pnl,4)),
            ('Funding (USDC)',    round(lt_fund,4), round(hl_fund,4), round(at_fund,4), round(total_fund,4)),
            ('P&L Total (USDC)', round(lt_pnl+lt_fund,4), round(hl_pnl+hl_fund,4), round(at_pnl+at_fund,4), round(total_pnl+total_fund,4)),
            ('Capital (USDC)',    round(lt_bal,2), round(hl_bal,2), round(at_bal,2), round(total_bal,2)),
            ('% Rentabilidad',   f"{round((lt_pnl+lt_fund)/lt_bal*100,2) if lt_bal else 0}%", f"{round((hl_pnl+hl_fund)/hl_bal*100,2) if hl_bal else 0}%", f"{round((at_pnl+at_fund)/at_bal*100,2) if at_bal else 0}%", f"{pct_rent}%"),
            ('Trades Cerrados',  len(lt_closes), len(hl_closes), len(at_closes), all_closes_n),
            ('Wins / Losses',    f"{lt_wins}/{len(lt_closes)-lt_wins}", f"{hl_wins}/{len(hl_closes)-hl_wins}", f"{at_wins}/{len(at_closes)-at_wins}", f"{all_wins_n}/{all_closes_n-all_wins_n}"),
            ('Win Rate',         f"{lt_wr}%", f"{hl_wr}%", f"{at_wr}%", f"{global_wr}%"),
        ]
        for i, row in enumerate(summary_data, 6):
            ws_gl.cell(i, 1, row[0]).font = Font(bold=True)
            for j, val in enumerate(row[1:], 2):
                cell = ws_gl.cell(i, j, val)
                cell.alignment = center; cell.border = thin
                style_pnl_cell(cell, val)

        # Diario combinado
        r = len(summary_data) + 8
        ws_gl.cell(r, 1, 'DIARIO COMBINADO').font = Font(bold=True, size=13, color='FFFFFF')
        ws_gl.cell(r, 1).fill = hdr_fill_gl
        ws_gl.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        r += 1
        daily_headers = ['Fecha', 'P&L Lighter', 'P&L Hyperliquid', 'P&L Aster', 'P&L Total Día', '% Total', 'Acumulado']
        style_header_row(ws_gl, r, daily_headers, hdr_fill_sub)
        r += 1
        dm_lt = build_daily_map(lt_closes)
        dm_hl = build_daily_map(hl_closes)
        dm_at = build_daily_map(at_closes)
        all_dates = sorted(set(list(dm_lt.keys())+list(dm_hl.keys())+list(dm_at.keys())), reverse=True)
        all_total_pnl_sum = total_pnl + total_fund
        cum = all_total_pnl_sum
        for d in all_dates:
            l = round(dm_lt.get(d,0),4); h = round(dm_hl.get(d,0),4)
            a = round(dm_at.get(d,0),4); tot = round(l+h+a,4)
            pct = round(tot/abs(all_total_pnl_sum)*100,1) if all_total_pnl_sum else 0
            ws_gl.cell(r,1,d); ws_gl.cell(r,1).alignment = center
            for col, val in [(2,l),(3,h),(4,a),(5,tot),(6,f"{pct}%"),(7,round(cum,4))]:
                cell = ws_gl.cell(r, col, val)
                cell.alignment = center; cell.border = thin
                style_pnl_cell(cell, val)
            cum -= tot; r += 1

        set_col_widths(ws_gl, [20, 16, 16, 16, 16, 12, 16])

        # ══════════════════════════════════════════
        # HOJA LIGHTER
        # ══════════════════════════════════════════
        ws_lt = wb.create_sheet('Lighter')
        ws_lt.cell(1,1,'LIGHTER').font = Font(bold=True,size=14,color='FFFFFF')
        ws_lt.cell(1,1).fill = hdr_fill_lt
        ws_lt.merge_cells('A1:H1')
        lt_sum_rows = [
            ('P&L Trades', round(lt_pnl,4)), ('Funding Total', round(lt_fund,4)),
            ('P&L Total', round(lt_pnl+lt_fund,4)), ('Capital Cuenta', round(lt_bal,2)),
            ('% Rentabilidad', f"{round((lt_pnl+lt_fund)/lt_bal*100,2) if lt_bal else 0}%"),
            ('Trades Cerrados', len(lt_closes)), ('Win Rate', f"{lt_wr}%"),
        ]
        for i,(k,v) in enumerate(lt_sum_rows, 2):
            ws_lt.cell(i,1,k).font = Font(bold=True)
            vc = ws_lt.cell(i,2,v); vc.border = thin; style_pnl_cell(vc,v)
        r = len(lt_sum_rows) + 4
        ws_lt.cell(r,1,'HISTORIAL DE TRADES').font = Font(bold=True,size=12,color='FFFFFF')
        ws_lt.cell(r,1).fill = hdr_fill_lt; ws_lt.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8); r+=1
        lt_t_hdrs = ['Fecha','Mercado','Tipo','Lado','Precio','Tamaño','P&L (USDC)','Fee']
        style_header_row(ws_lt, r, lt_t_hdrs, hdr_fill_lt); r+=1
        for t in sorted(trades.values(), key=lambda t: int(t.get('ts',0) or 0), reverse=True):
            row_data = [fmt_ts(t.get('ts')), t.get('symbol',''),
                'Cierre' if t.get('tradeType')=='close' else 'Apertura',
                (t.get('side','') or '').upper(),
                round(float(t.get('price') or 0),4) or '',
                round(float(t.get('size') or 0),6) or '',
                t.get('pnl') if t.get('pnl') is not None else '',
                round(float(t.get('fee') or 0),6) or '']
            for col,val in enumerate(row_data,1):
                cell=ws_lt.cell(r,col,val); cell.border=thin
                if col==1: cell.alignment=center
            style_pnl_cell(ws_lt.cell(r,7), row_data[6]); r+=1
        r+=1
        ws_lt.cell(r,1,'FUNDING FEES').font = Font(bold=True,size=12,color='FFFFFF')
        ws_lt.cell(r,1).fill = hdr_fill_lt; ws_lt.merge_cells(start_row=r,start_column=1,end_row=r,end_column=5); r+=1
        style_header_row(ws_lt, r, ['Fecha','Mercado','Lado','Pago (USDC)','Tasa'], hdr_fill_lt); r+=1
        for f in sorted(funding.values(), key=lambda f: int(f.get('ts',0) or 0), reverse=True):
            row_data=[fmt_ts(f.get('ts')),f.get('symbol',''),f.get('side',''),f.get('payment',''),f.get('rate','')]
            for col,val in enumerate(row_data,1):
                cell=ws_lt.cell(r,col,val); cell.border=thin
            style_pnl_cell(ws_lt.cell(r,4), f.get('payment')); r+=1
        set_col_widths(ws_lt,[18,12,10,8,14,14,16,14])

        # ══════════════════════════════════════════
        # HOJA HYPERLIQUID
        # ══════════════════════════════════════════
        ws_hl = wb.create_sheet('Hyperliquid')
        ws_hl.cell(1,1,'HYPERLIQUID').font = Font(bold=True,size=14,color='FFFFFF')
        ws_hl.cell(1,1).fill = hdr_fill_hl; ws_hl.merge_cells('A1:H1')
        hl_sum_rows = [
            ('P&L Trades', round(hl_pnl,4)), ('Funding Total', round(hl_fund,4)),
            ('P&L Total', round(hl_pnl+hl_fund,4)), ('Capital Cuenta', round(hl_bal,2)),
            ('% Rentabilidad', f"{round((hl_pnl+hl_fund)/hl_bal*100,2) if hl_bal else 0}%"),
            ('Trades Cerrados', len(hl_closes)), ('Win Rate', f"{hl_wr}%"),
        ]
        for i,(k,v) in enumerate(hl_sum_rows,2):
            ws_hl.cell(i,1,k).font=Font(bold=True)
            vc=ws_hl.cell(i,2,v); vc.border=thin; style_pnl_cell(vc,v)
        r=len(hl_sum_rows)+4
        ws_hl.cell(r,1,'HISTORIAL DE TRADES').font=Font(bold=True,size=12,color='FFFFFF')
        ws_hl.cell(r,1).fill=hdr_fill_hl; ws_hl.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8); r+=1
        style_header_row(ws_hl,r,['Fecha','Mercado','Tipo','Lado','Precio','Tamaño','P&L (USDC)','Fee'],hdr_fill_hl); r+=1
        for t in sorted(hl_trades.values(), key=lambda t: int(t.get('ts',0) or 0), reverse=True):
            row_data=[fmt_ts(t.get('ts')),t.get('symbol',''),'Cierre' if t.get('tradeType')=='close' else 'Apertura',(t.get('side','') or '').upper(),round(float(t.get('price') or 0),4) or '',round(float(t.get('size') or 0),6) or '',t.get('pnl') if t.get('pnl') is not None else '',round(float(t.get('fee') or 0),6) or '']
            for col,val in enumerate(row_data,1):
                cell=ws_hl.cell(r,col,val); cell.border=thin
                if col==1: cell.alignment=center
            style_pnl_cell(ws_hl.cell(r,7),row_data[6]); r+=1
        r+=1
        ws_hl.cell(r,1,'FUNDING FEES').font=Font(bold=True,size=12,color='FFFFFF')
        ws_hl.cell(r,1).fill=hdr_fill_hl; ws_hl.merge_cells(start_row=r,start_column=1,end_row=r,end_column=3); r+=1
        style_header_row(ws_hl,r,['Fecha','Mercado','Pago (USDC)'],hdr_fill_hl); r+=1
        for f in sorted(hl_funding.values(), key=lambda f: int(f.get('ts',0) or 0), reverse=True):
            row_data=[fmt_ts(f.get('ts')),f.get('symbol',''),f.get('payment','')]
            for col,val in enumerate(row_data,1):
                cell=ws_hl.cell(r,col,val); cell.border=thin
            style_pnl_cell(ws_hl.cell(r,3),f.get('payment')); r+=1
        set_col_widths(ws_hl,[18,12,10,8,14,14,16,14])

        # ══════════════════════════════════════════
        # HOJA ASTER
        # ══════════════════════════════════════════
        ws_at = wb.create_sheet('Aster')
        ws_at.cell(1,1,'ASTER').font=Font(bold=True,size=14,color='FFFFFF')
        ws_at.cell(1,1).fill=hdr_fill_at; ws_at.merge_cells('A1:H1')
        at_sum_rows = [
            ('P&L Trades', round(at_pnl,4)), ('Funding Total', round(at_fund,4)),
            ('P&L Total', round(at_pnl+at_fund,4)), ('Capital Cuenta', round(at_bal,2)),
            ('% Rentabilidad', f"{round((at_pnl+at_fund)/at_bal*100,2) if at_bal else 0}%"),
            ('Trades Cerrados', len(at_closes)), ('Win Rate', f"{at_wr}%"),
        ]
        for i,(k,v) in enumerate(at_sum_rows,2):
            ws_at.cell(i,1,k).font=Font(bold=True)
            vc=ws_at.cell(i,2,v); vc.border=thin; style_pnl_cell(vc,v)
        r=len(at_sum_rows)+4
        ws_at.cell(r,1,'HISTORIAL DE TRADES').font=Font(bold=True,size=12,color='FFFFFF')
        ws_at.cell(r,1).fill=hdr_fill_at; ws_at.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8); r+=1
        style_header_row(ws_at,r,['Fecha','Mercado','Tipo','Lado','Precio','Tamaño','P&L (USDC)','Fee'],hdr_fill_at); r+=1
        for t in sorted(aster_trades.values(), key=lambda t: int(t.get('ts',0) or 0), reverse=True):
            row_data=[fmt_ts(t.get('ts')),t.get('symbol',''),'Cierre',(t.get('side','') or '').upper(),round(float(t.get('price') or 0),4) or '',round(float(t.get('size') or 0),6) or '',t.get('pnl') if t.get('pnl') is not None else '',round(float(t.get('fee') or 0),6) or '']
            for col,val in enumerate(row_data,1):
                cell=ws_at.cell(r,col,val); cell.border=thin
                if col==1: cell.alignment=center
            style_pnl_cell(ws_at.cell(r,7),row_data[6]); r+=1
        r+=1
        ws_at.cell(r,1,'FUNDING FEES').font=Font(bold=True,size=12,color='FFFFFF')
        ws_at.cell(r,1).fill=hdr_fill_at; ws_at.merge_cells(start_row=r,start_column=1,end_row=r,end_column=3); r+=1
        style_header_row(ws_at,r,['Fecha','Mercado','Pago (USDC)'],hdr_fill_at); r+=1
        for f in sorted(aster_funding.values(), key=lambda f: int(f.get('ts',0) or 0), reverse=True):
            row_data=[fmt_ts(f.get('ts')),f.get('symbol',''),f.get('payment','')]
            for col,val in enumerate(row_data,1):
                cell=ws_at.cell(r,col,val); cell.border=thin
            style_pnl_cell(ws_at.cell(r,3),f.get('payment')); r+=1
        set_col_widths(ws_at,[18,12,10,8,14,14,16,14])

        # Save
        buf = _io.BytesIO()
        wb.save(buf); buf.seek(0)
        fname = f"registro_dex_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M')}.xlsx"
        return web.Response(
            body=buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{fname}"',
                     'Access-Control-Allow-Origin': '*'}
        )
    except Exception as e:
        log.error(f"Excel export error: {e}")
        import traceback; traceback.print_exc()
        return cors(web.json_response({'error': str(e)}, status=500))


async def h_options(req):
    return web.Response(headers={
        'Access-Control-Allow-Origin': '*',
        'Access-Control-Allow-Methods': 'GET, OPTIONS',
        'Access-Control-Allow-Headers': '*'
    })

async def on_start(app):
    app['task'] = asyncio.ensure_future(ws_listener())
    if HL_WALLET:
        app['hl_task'] = asyncio.ensure_future(hl_main_loop())
    if ASTER_API_KEY:
        app['aster_task'] = asyncio.ensure_future(aster_main_loop())

async def aster_main_loop():
    await load_aster_data()
    while True:
        await asyncio.sleep(900)
        await aster_incremental()

async def hl_main_loop():
    await load_hl_data()
    while True:
        await asyncio.sleep(900)  # refresh every 15 min
        await hl_incremental()

async def on_stop(app):
    app['task'].cancel()
    try: await app['task']
    except asyncio.CancelledError: pass
    if 'hl_task' in app:
        app['hl_task'].cancel()
        try: await app['hl_task']
        except asyncio.CancelledError: pass
    if 'aster_task' in app:
        app['aster_task'].cancel()
        try: await app['aster_task']
        except asyncio.CancelledError: pass

def create_app():
    app = web.Application()
    app.router.add_get('/', h_root)
    app.router.add_get('/status', h_status)
    app.router.add_get('/trades', h_trades)
    app.router.add_get('/funding', h_funding)
    app.router.add_get('/positions', h_positions)
    app.router.add_get('/summary', h_summary)
    app.router.add_get('/debug/account', h_account_debug)
    app.router.add_get('/hl/summary', h_hl_summary)
    app.router.add_get('/aster/summary', h_aster_summary)
    app.router.add_get('/aster/trades', h_aster_trades)
    app.router.add_get('/aster/positions', h_aster_positions)
    app.router.add_post('/aster/upload_csv', h_aster_upload_csv)
    app.router.add_post('/aster/clear_csv', h_aster_clear_csv)
    app.router.add_get('/hl/debug', h_hl_debug)
    app.router.add_get('/hl/funding_test', h_hl_funding_test)
    app.router.add_get('/hl/ledger', h_hl_ledger_inspect)
    app.router.add_post('/hl/upload_funding', h_hl_upload_funding)
    app.router.add_post('/hl/clear_funding', h_hl_clear_funding)
    app.router.add_get('/hl/trades', h_hl_trades)
    app.router.add_get('/hl/positions', h_hl_positions)
    app.router.add_get('/export/excel', h_export_excel)
    app.router.add_get('/eur/rates', h_eur_rates)
    app.router.add_options('/{p:.*}', h_options)
    app.on_startup.append(on_start)
    app.on_cleanup.append(on_stop)
    return app

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 10000))
    log.info(f"Starting on port {port}")
    web.run_app(create_app(), port=port)
